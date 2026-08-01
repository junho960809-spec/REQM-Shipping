from pathlib import Path

from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADERS = [
    "주문번호", "판매처", "상품명", "수량", "수령자", "핸드폰", "우편번호",
    "주소", "배송메세지", "송장번호", "일련번호",
]


def export_wekep(orders: list[dict[str, str]], file_path: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "택배출고"
    sheet.append(HEADERS)
    for order in orders:
        product_name = order.get("matched_product") or order.get("matched_name") or order.get("product_name")
        sheet.append(
            [
                order.get("order_number", ""),
                order.get("channel", ""),
                product_name,
                order.get("quantity", ""),
                order.get("recipient", ""),
                order.get("phone", ""),
                order.get("zipcode", ""),
                order.get("address", ""),
                order.get("message", ""),
                "",
                order.get("serial_number", ""),
            ]
        )

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = [22, 14, 45, 9, 14, 18, 11, 55, 35, 18, 14]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    workbook.save(Path(file_path))


def export_with_format(orders: list[dict[str, str]], file_path: str, profile: dict) -> None:
    if profile.get("id") == "default_b2c":
        export_wekep(orders, file_path)
        return

    template_path = str(profile.get("template_path", ""))
    if template_path:
        workbook = load_workbook(template_path)
        sheet = workbook.active
        header_row = int(profile.get("header_row", 0)) + 1
        headers = [cell.value for cell in sheet[header_row]]
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "B2C 사입형"
        headers = list(profile.get("headers") or [])
        sheet.append(headers)
        header_row = 1
        fill = PatternFill("solid", fgColor="FFA395")
        for index in (3, 5, 8, 10, 11, 12):
            sheet.cell(header_row, index).fill = fill
        for cell in sheet[header_row]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    header_positions = {
        str(value).strip(): index for index, value in enumerate(headers, start=1) if value is not None
    }
    mapping = dict(profile.get("mapping") or {})
    missing_headers = [header for header in mapping.values() if header not in header_positions]
    if missing_headers:
        workbook.close()
        raise ValueError("출력 양식에서 연결된 열을 찾지 못했습니다: " + ", ".join(missing_headers))

    start_row = header_row + 1
    template_style_row = start_row
    if template_path:
        for row_index in range(start_row, sheet.max_row + 1):
            for header in mapping.values():
                sheet.cell(row_index, header_positions[header]).value = None
    for row_index, order in enumerate(orders, start=start_row):
        if row_index > start_row and template_path:
            for column in range(1, max(len(headers), 1) + 1):
                source = sheet.cell(template_style_row, column)
                target = sheet.cell(row_index, column)
                if source.has_style:
                    target._style = copy(source._style)
                if source.number_format:
                    target.number_format = source.number_format
        for key, header in mapping.items():
            value = order.get(key, "")
            if key == "product_name":
                value = order.get("matched_product") or order.get("matched_name") or value
            sheet.cell(row_index, header_positions[header]).value = value

    workbook.save(Path(file_path))
    workbook.close()

