from pathlib import Path
import re
from typing import Any

import openpyxl
import pdfplumber


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def norm(value: Any) -> str:
    return "".join(clean(value).lower().split()).replace(".", "")


COLOR_WORDS = {
    "블랙", "화이트", "핑크", "그레이", "그린", "블루", "레드", "옐로우", "오렌지",
    "퍼플", "베이지", "민트", "라벤더", "라밴더", "버터", "캐롯", "캐럿", "샌드",
    "코발트블루", "세이지민트", "올리브", "브라운", "네이비",
}


def _model_tokens(value: Any) -> set[str]:
    text = clean(value).lower()
    return {
        re.sub(r"[^a-z0-9]", "", token)
        for token in re.findall(r"[a-z]{1,8}[-_ ]?\d+[a-z0-9]*", text)
        if len(re.sub(r"[^a-z0-9]", "", token)) >= 4
    }


def _colors(value: Any) -> set[str]:
    text = norm(value).replace("블루투스", "")
    aliases = {"라밴더": "라벤더", "캐럿": "캐롯"}
    return {aliases.get(color, color) for color in COLOR_WORDS if color in text}


def barcode_name_error(product_name: str, item: dict) -> str:
    """Return a reason when a barcode's DB item clearly conflicts with the source product."""
    db_text = " ".join(
        clean(item.get(key, ""))
        for key in ("standard_name", "model", "color", "form")
    )
    source_norm, db_norm = norm(product_name), norm(db_text)

    source_models = _model_tokens(product_name)
    db_models = _model_tokens(db_text)
    model_overlap = any(
        source.startswith(db) or db.startswith(source)
        for source in source_models for db in db_models
    )
    if source_models and db_models and not model_overlap:
        return f"상품 모델 불일치: 파일 {', '.join(sorted(source_models))} / DB {', '.join(sorted(db_models))}"

    source_colors = _colors(product_name)
    db_colors = _colors(db_text)
    if source_colors and db_colors and source_colors.isdisjoint(db_colors):
        return f"상품 색상 불일치: 파일 {', '.join(sorted(source_colors))} / DB {', '.join(sorted(db_colors))}"

    source_is_set = "세트" in source_norm or "+" in clean(product_name)
    db_is_set = "세트" in db_norm or str(item.get("item_code", "")).upper().startswith("SET-")
    if source_is_set and not db_is_set:
        return "파일은 세트 상품이지만 바코드는 DB 단품에 연결됨"
    return ""


PRODUCT_HEADERS = {"상품명", "품명", "제품명", "상품", "itemname", "productname"}
QUANTITY_HEADERS = {"수량", "입고수량", "발주수량", "qty", "quantity"}
REF_HEADERS = {"refno", "바코드"}
SKU_HEADERS = {"skuno", "sku", "상품코드", "품목코드"}


def _header_index(row: list[Any], candidates: set[str]) -> int | None:
    for index, value in enumerate(row):
        key = norm(value)
        if key in candidates:
            return index
    return None


def _quantity(value: Any) -> str:
    text = clean(value).replace(",", "")
    try:
        number = float(text)
    except (TypeError, ValueError):
        return ""
    if number <= 0:
        return ""
    return str(int(number)) if number.is_integer() else str(number)


def _detect_channel(text: str) -> str:
    compact_text = norm(text)
    for keyword, channel in (
        ("롯데", "롯데면세점"),
        ("신라", "신라면세점"),
        ("신세계", "신세계면세점"),
        ("현대", "현대면세점"),
        ("시티", "시티면세점"),
    ):
        if keyword in compact_text:
            return channel
    return "면세점"


def _simple_orders_from_rows(
    rows: list[list[Any]], source_label: str, channel: str
) -> list[dict[str, str]]:
    """Read only product name and quantity from a fixed-layout table."""
    for header_index, row in enumerate(rows[:40]):
        product_col = _header_index(row, PRODUCT_HEADERS)
        quantity_col = _header_index(row, QUANTITY_HEADERS)
        if product_col is None or quantity_col is None:
            continue
        ref_col = _header_index(row, REF_HEADERS)
        sku_col = _header_index(row, SKU_HEADERS)
        result: list[dict[str, str]] = []
        for row_offset, data in enumerate(rows[header_index + 1 :], start=header_index + 2):
            product = clean(data[product_col]) if product_col < len(data) else ""
            quantity = _quantity(data[quantity_col]) if quantity_col < len(data) else ""
            if not product or not quantity:
                continue
            ref_no = clean(data[ref_col]) if ref_col is not None and ref_col < len(data) else ""
            sku_no = clean(data[sku_col]) if sku_col is not None and sku_col < len(data) else ""
            result.append(
                {
                    "source_row": f"{source_label} · {row_offset}",
                    "order_number": "",
                    "channel": channel,
                    "product_name": re.sub(r"\s+", " ", product).strip(),
                    "source_item_code": ref_no or sku_no,
                    "ref_no": ref_no,
                    "sku_no": sku_no,
                    "options": "",
                    "quantity": quantity,
                    "recipient": "",
                    "phone": "",
                    "zipcode": "",
                    "address": "",
                    "message": "",
                    "matched_name": product,
                    "match_method": "name_or_code",
                }
            )
        if result:
            return result
    return []


def load_simple_duty_free(file_path: str) -> tuple[list[dict[str, str]], str]:
    """Load a sparse duty-free document using its fixed table geometry.

    Only product name and quantity are required in the source document. The UI
    applies a saved destination after loading.
    """
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        orders: list[dict[str, str]] = []
        document_text: list[str] = [path.stem]
        with pdfplumber.open(path) as document:
            for page_number, page in enumerate(document.pages, start=1):
                page_text = page.extract_text() or ""
                document_text.append(page_text)
                channel = _detect_channel(" ".join(document_text))
                for table_number, table in enumerate(page.extract_tables() or [], start=1):
                    parsed = _simple_orders_from_rows(
                        table or [], f"PDF {page_number}쪽 표 {table_number}", channel
                    )
                    if parsed:
                        orders.extend(parsed)
                        break
        if not orders:
            raise ValueError("PDF에서 상품명과 수량이 있는 고정 표를 찾지 못했습니다.")
        channel = _detect_channel(" ".join(document_text))
        for order in orders:
            order["channel"] = channel
        return orders, channel

    if suffix == ".xlsx":
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()
    elif suffix == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(path)
        sheet = workbook.sheet_by_index(0)
        rows = [sheet.row_values(index) for index in range(sheet.nrows)]
    else:
        raise ValueError("면세점 출고 입력은 PDF, XLSX, XLS 파일을 지원합니다.")

    channel = _detect_channel(path.stem + " " + " ".join(clean(value) for row in rows[:10] for value in row))
    orders = _simple_orders_from_rows(rows, "Excel", channel)
    if not orders:
        raise ValueError("Excel에서 상품명과 수량 열을 찾지 못했습니다.")
    return orders, channel


def load_duty_free(file_path: str) -> tuple[list[dict[str, str]], str] | None:
    path = Path(file_path)
    if path.suffix.lower() != ".xlsx":
        return None
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = book.active
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    file_text = " ".join(clean(value) for row in rows[:8] for value in row if value)

    for header_index, row in enumerate(rows[:15]):
        headers = {norm(value): index for index, value in enumerate(row) if value is not None}
        tm_code = headers.get("tm상품코드")
        tm_product = headers.get("tm상품명")
        tm_qty = headers.get("발주수량")
        tm_store = headers.get("매장명")
        if tm_code is not None and tm_product is not None and tm_qty is not None and tm_store is not None:
            order_date_match = re.search(r"발\s*주\s*일\s*:\s*([0-9.\\/-]+)", file_text)
            order_date = order_date_match.group(1) if order_date_match else ""
            store_details: dict[str, dict[str, str]] = {}
            for info_row in rows[header_index + 1 :]:
                first_text = next((clean(value) for value in info_row if clean(value)), "")
                if not first_text.startswith("＊"):
                    continue
                store_label = first_text.lstrip("＊").strip()
                values = [clean(value) for value in info_row if clean(value)]
                address = next(
                    (value for value in values if "시 " in value or "도 " in value or "공항로" in value or "터미널대로" in value),
                    "",
                )
                phone = next((value for value in values if re.search(r"\d{2,3}-\d{3,4}-\d{4}", value)), "")
                manager_match = re.search(r"([가-힣]{2,5}\s*매니저)", address)
                recipient = manager_match.group(1).replace(" ", "") if manager_match else store_label
                store_details[norm(store_label)] = {
                    "recipient": recipient,
                    "address": address,
                    "phone": phone,
                    "store_label": store_label,
                }

            result = []
            for source_row, data in enumerate(rows[header_index + 1 :], start=header_index + 2):
                product = clean(data[tm_product]) if tm_product < len(data) else ""
                external_code = clean(data[tm_code]) if tm_code < len(data) else ""
                quantity = clean(data[tm_qty]) if tm_qty < len(data) else ""
                store = clean(data[tm_store]) if tm_store < len(data) else ""
                try:
                    numeric_quantity = float(quantity.replace(",", ""))
                except ValueError:
                    numeric_quantity = 0
                if not product or numeric_quantity <= 0:
                    continue
                details = next(
                    (
                        value for key, value in store_details.items()
                        if norm(store) and norm(store) in key
                    ),
                    {},
                )
                store_label = details.get("store_label") or store or "매장"
                result.append({
                    "source_row": str(source_row),
                    "order_number": " ".join(filter(None, ["트래블메이트", store_label, order_date])),
                    "channel": "트래블메이트",
                    "product_name": product,
                    "source_item_code": external_code,
                    "options": "",
                    "quantity": str(int(numeric_quantity)) if numeric_quantity.is_integer() else str(numeric_quantity),
                    "recipient": details.get("recipient", store_label),
                    "phone": details.get("phone", ""),
                    "zipcode": "",
                    "address": details.get("address", ""),
                    "message": f"매장명: {store_label}",
                    "matched_name": re.sub(r"^\s*\[리큐엠\]\s*", "", product).strip(),
                    "external_code": external_code,
                    "match_method": "name_or_code",
                    "embedded_destination": bool(details.get("address")),
                })
            if result:
                destination = result[0].get("message", "").replace("매장명: ", "")
                return result, f"트래블메이트 {destination} 발주서"

        city_barcode = headers.get("바코드")
        city_product = headers.get("상품명")
        city_qty = headers.get("수량")
        if city_barcode is not None and city_product is not None and city_qty is not None and "box" in headers:
            store_code = "403" if "403" in file_text or "403" in path.name else "606" if "606" in file_text or "606" in path.name else ""
            store_name = f"T2 {store_code}매장" if store_code else "시티면세점 매장"
            result = []
            for source_row, data in enumerate(rows[header_index + 1 :], start=header_index + 2):
                barcode = clean(data[city_barcode]) if city_barcode < len(data) else ""
                product = clean(data[city_product]) if city_product < len(data) else ""
                if not barcode and not product:
                    continue
                get = lambda name: clean(data[headers[name]]) if name in headers and headers[name] < len(data) else ""
                result.append({
                    "source_row": str(source_row), "order_number": store_name, "channel": "시티면세점",
                    "product_name": product, "options": "", "quantity": clean(data[city_qty]),
                    "recipient": get("수령인"), "phone": get("연락처"), "zipcode": "",
                    "address": get("주소"), "message": f"BOX {get('box')}" if get("box") else "",
                    "matched_name": barcode, "barcode": barcode, "store_code": store_code,
                })
            return result, f"시티면세점 {store_name}"

        ref_col = next((index for key, index in headers.items() if key in {"refno", "바코드"}), None)
        product_col = headers.get("상품명")
        qty_col = headers.get("수량")
        code_col = headers.get("skuno") if "skuno" in headers else headers.get("상품코드")
        if ref_col is not None and product_col is not None and qty_col is not None:
            source_text = file_text + " " + path.name
            duty_name = (
                "롯데면세점" if "롯데" in source_text else
                "현대면세점" if "현대" in source_text else
                "신라면세점" if "신라" in source_text else
                "신세계면세점" if "신세계" in source_text else
                "시티면세점" if "시티" in source_text or "넥서스코프" in source_text else
                "면세점(종류 확인 필요)"
            )
            result = []
            for source_row, data in enumerate(rows[header_index + 1 :], start=header_index + 2):
                barcode = clean(data[ref_col]) if ref_col < len(data) else ""
                product = clean(data[product_col]) if product_col < len(data) else ""
                if not barcode and not product:
                    continue
                external_code = clean(data[code_col]) if code_col is not None and code_col < len(data) else ""
                result.append({
                    "source_row": str(source_row), "order_number": external_code, "channel": duty_name,
                    "product_name": product, "options": "", "quantity": clean(data[qty_col]),
                    "recipient": "", "phone": "", "zipcode": "", "address": "", "message": "",
                    "matched_name": barcode, "barcode": barcode, "external_code": external_code,
                })
            return result, duty_name
    return None


def match_barcodes(rows: list[dict[str, str]], barcodes: list[dict], items: list[dict]) -> None:
    barcode_map = {str(row.get("barcode", "")): str(row.get("item_code", "")) for row in barcodes if row.get("is_active", True)}
    item_map = {str(row.get("item_code", "")): row for row in items}
    for row in rows:
        item_code = barcode_map.get(row.get("barcode", ""), "")
        item = item_map.get(item_code)
        if item:
            mismatch = barcode_name_error(row.get("product_name", ""), item)
            if mismatch:
                row.update({
                    "status": "barcode_error", "matched_product": str(item.get("standard_name", "")),
                    "components": item_code, "reason": "바코드-상품 불일치 · " + mismatch,
                })
            else:
                row.update({
                    "status": "exact", "matched_product": str(item.get("standard_name", "")),
                    "components": item_code, "reason": "면세점 바코드 정확 일치",
                })
        else:
            row.update({
                "status": "barcode_error", "matched_product": "", "components": "",
                "reason": "바코드 오류 · DB에 등록되지 않은 바코드",
            })

