from __future__ import annotations

from datetime import date, datetime, timedelta
import os
from pathlib import Path
import tempfile

from openpyxl import load_workbook
from weekly_inventory_supabase import replace_period


ERP_HASH = (
    "menuType=MENUTREE_000004&menuSeq=MENUTREE_000030&"
    "groupSeq=MENUTREE_000030&prgId=C000030&depth=2"
)
ERP_URL = f"https://loginab.ecount.com/ec5/view/erp?w_flag=1#{ERP_HASH}"
WEEKDAYS = "월화수목금토일"


def previous_inventory_week(reference: date | None = None) -> tuple[date, date]:
    current = reference or date.today()
    days_since_thursday = (current.weekday() - 3) % 7
    end = current - timedelta(days=days_since_thursday)
    return end - timedelta(days=6), end


def parse_ecount_sales_excel(path: str | Path, start_date: date, end_date: date) -> list[list]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        result = []
        for row in sheet.iter_rows(min_row=3, values_only=True):
            raw_date = str(row[0] or "").strip()
            if len(raw_date) != 8 or not raw_date.isdigit() or len(row) < 9:
                continue
            sale_date = datetime.strptime(raw_date, "%Y%m%d").date()
            if not start_date <= sale_date <= end_date:
                continue
            item_name = str(row[4] or "").strip()
            category = item_name[1:item_name.find("]")] if item_name.startswith("[") and "]" in item_name else ""
            iso_week = sale_date.isocalendar().week
            result.append([
                sale_date.year, sale_date.month, sale_date.day, WEEKDAYS[sale_date.weekday()],
                f"{iso_week}W", f"{start_date:%m/%d}~{end_date:%m/%d}", category,
                int(sale_date.strftime("%Y%m%d")), str(row[1] or "").strip(), str(row[2] or "").strip(),
                str(row[3] or "").strip(), item_name, _number(row[5]), _integer(row[6]),
                _integer(row[7]), _integer(row[8]),
            ])
        if not result:
            raise ValueError("이카운트 판매현황 Excel에서 상세 판매행을 찾지 못했습니다.")
        return result
    finally:
        workbook.close()


def _number(value) -> float:
    try:
        return float(str(value or 0).replace(",", ""))
    except (TypeError, ValueError):
        return 0.0


def _integer(value) -> int:
    return int(round(_number(value)))


def _select_dropdown(page, button, value: str) -> None:
    if value in button.inner_text().strip():
        return
    button.click()
    page.get_by_role("link", name=value, exact=True).click()


def download_ecount_sales(credentials: dict, start_date: date, end_date: date) -> Path:
    from playwright.sync_api import sync_playwright

    profile = Path(os.getenv("LOCALAPPDATA", str(Path.home()))) / "REQM" / "ecount_sales_browser"
    profile.mkdir(parents=True, exist_ok=True)
    download_dir = Path(tempfile.mkdtemp(prefix="reqm-ecount-sales-"))
    with sync_playwright() as playwright:
        context = playwright.chromium.launch_persistent_context(
            str(profile), channel="chrome", headless=True, accept_downloads=True,
            downloads_path=str(download_dir), viewport={"width": 1500, "height": 1000},
        )
        try:
            page = context.pages[0] if context.pages else context.new_page()
            page.goto(ERP_URL, wait_until="domcontentloaded", timeout=60000)
            page.wait_for_timeout(2000)
            if page.locator("#com_code").count():
                page.locator("#com_code").fill(str(credentials.get("company_code") or "304293"))
                page.locator("#id").fill(str(credentials.get("user_id") or ""))
                page.locator("#passwd").fill(str(credentials.get("password") or ""))
                page.locator("#save").click()
                page.wait_for_timeout(3000)
                if page.locator("#com_code").count():
                    raise RuntimeError("이카운트 로그인에 실패했습니다. 연동 계정의 아이디와 비밀번호를 확인하세요.")
                session_url = page.url.split("#", 1)[0]
                page.goto(f"{session_url}#{ERP_HASH}", wait_until="domcontentloaded", timeout=60000)
                page.wait_for_timeout(2000)
            visible_date_box = page.locator(
                '[data-item-key="data_dt_salesXsum_status_search"]:visible, '
                '[data-item-key="data_dt_salesXstatus_search"]:visible'
            )
            if not visible_date_box.count():
                sales_status = page.get_by_text("판매현황", exact=True)
                if not sales_status.count():
                    raise RuntimeError("이카운트 메뉴에서 판매현황을 찾지 못했습니다.")
                sales_status.first.click()
                page.wait_for_timeout(2500)
            visible_date_box.last.wait_for(timeout=60000)
            page.get_by_role("radio", name="집계", exact=True).check()
            page.wait_for_timeout(500)
            format_button = page.locator("button").filter(has_text="품목별 [규격]")
            if not format_button.count():
                sort_button = page.get_by_role("button", name="정렬기준 설정", exact=True)
                sort_button.locator("xpath=preceding::button[1]").click()
                page.get_by_role("link", name="품목별 [규격]", exact=True).click()
            condition_fields = []
            for label, value in (("집계조건1", "일별"), ("집계조건2", "거래처별"), ("집계조건3", "품목별")):
                field = page.get_by_role("textbox", name=label, exact=True)
                field.fill(value)
                condition_fields.append(field)
            condition_fields[0].locator("xpath=preceding::button[1]").click()
            page.wait_for_timeout(300)
            close_buttons = page.get_by_role("button", name="닫기", exact=True)
            if close_buttons.count():
                close_buttons.last.click()
            date_box = page.locator(
                '[data-item-key="data_dt_salesXsum_status_search"], [data-item-key="data_dt_salesXstatus_search"]'
            ).last
            year_buttons = date_box.locator('button[data-id="year"]')
            if year_buttons.count() >= 2:
                _select_dropdown(page, year_buttons.nth(0), str(start_date.year))
                _select_dropdown(page, year_buttons.nth(1), str(end_date.year))
            month_buttons = date_box.locator('button[data-id="month"]')
            _select_dropdown(page, month_buttons.nth(0), f"{start_date.month:02d}")
            _select_dropdown(page, month_buttons.nth(1), f"{end_date.month:02d}")
            day_inputs = date_box.get_by_role("textbox")
            day_inputs.nth(0).fill(f"{start_date.day:02d}")
            day_inputs.nth(1).fill(f"{end_date.day:02d}")
            page.get_by_role("button", name="검색(F8)", exact=True).click()
            page.get_by_text("총합계", exact=True).wait_for(timeout=60000)
            excel_buttons = page.get_by_role("button", name="", exact=True)
            excel_buttons.nth(1).click()
            with page.expect_download(timeout=60000) as download_info:
                page.get_by_role("link", name="Excel(데이터)", exact=True).click()
            download = download_info.value
            target = download_dir / (download.suggested_filename or "ecount_sales.xlsx")
            download.save_as(str(target))
            return target
        finally:
            context.close()


def sync_ecount_sales(client, credentials: dict, start_date: date | None = None, end_date: date | None = None) -> dict:
    calculated_start, calculated_end = previous_inventory_week()
    start_date = start_date or calculated_start
    end_date = end_date or calculated_end
    downloaded = download_ecount_sales(credentials, start_date, end_date)
    rows = parse_ecount_sales_excel(downloaded, start_date, end_date)
    result = replace_period(client, start_date, end_date, rows)
    result.update({"start_date": start_date, "end_date": end_date, "downloaded": str(downloaded)})
    return result
