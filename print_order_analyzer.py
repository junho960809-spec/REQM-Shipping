from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
import subprocess
import sys

import pdfplumber
from openpyxl import load_workbook


IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"}
FIELD_ALIASES = {
    "item_code": ("품목코드", "상품코드", "제품코드", "모델코드"),
    "recipient": ("수령인", "성명", "받는사람", "받는 사람", "배송처", "수취인"),
    "contact": ("연락처", "휴대폰", "핸드폰", "전화번호", "TEL", "전화"),
    "address": ("배송지", "배송주소", "받는곳", "받는 곳", "주소"),
    "request_date": ("납기", "납기일", "출고일", "출고요청일", "발송일"),
    "product": ("품명", "상품명", "제품명", "데이터명"),
    "quantity": ("수량", "주문수량", "발주수량", "총수량"),
    "printing": ("인쇄내용", "인쇄 요청", "인쇄요청", "사용"),
    "packaging": ("포장", "포장선택", "선물포장"),
    "delivery": ("배송", "배송방법", "발송방법"),
}


@dataclass
class AnalysisResult:
    source_type: str
    vendor: str
    fields: dict[str, str]
    confidence: dict[str, int]
    raw_text: str


def _ocr_script_path() -> Path:
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    return base / "assets" / "windows_ocr.ps1"


def extract_text(path: str | Path) -> tuple[str, str]:
    source = Path(path)
    suffix = source.suffix.casefold()
    if suffix in IMAGE_EXTENSIONS:
        script = _ocr_script_path()
        if not script.exists():
            raise FileNotFoundError("Windows OCR 구성파일을 찾지 못했습니다.")
        completed = subprocess.run(
            ["powershell.exe", "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", str(script), "-Path", str(source)],
            capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=60,
        )
        if completed.returncode != 0:
            raise RuntimeError(completed.stderr.strip() or "이미지 OCR에 실패했습니다.")
        return completed.stdout.strip(), "이미지 OCR"
    if suffix == ".pdf":
        with pdfplumber.open(source) as document:
            return "\n".join(page.extract_text() or "" for page in document.pages), "PDF"
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(source, read_only=True, data_only=True)
        try:
            lines = []
            for sheet in workbook.worksheets:
                lines.append(f"[시트:{sheet.title}]")
                for row in sheet.iter_rows(values_only=True):
                    values = [str(value).strip() for value in row if value not in (None, "")]
                    if values:
                        lines.append(" | ".join(values))
            return "\n".join(lines), "Excel"
        finally:
            workbook.close()
    if suffix in {".txt", ".csv"}:
        return source.read_text(encoding="utf-8-sig", errors="replace"), "텍스트"
    raise ValueError("지원 형식: PNG/JPG/BMP/TIFF, PDF, XLSX/XLSM, TXT/CSV")


def _clean(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip(" |:-\t")


def _after_alias(text: str, aliases: tuple[str, ...], limit: int = 80) -> str:
    for alias in aliases:
        match = re.search(rf"{re.escape(alias)}\s*[:：]?\s*([^\n|]{{2,{limit}}})", text, re.IGNORECASE)
        if match:
            return _clean(match.group(1))
    return ""


def analyze_text(text: str, source_type: str = "텍스트") -> AnalysisResult:
    normalized = text.replace("\r", "\n")
    vendor = "고려기프트" if "고려기프트" in normalized or "고켴기프" in normalized else "신규 업체"
    fields = {key: _after_alias(normalized, aliases) for key, aliases in FIELD_ALIASES.items()}

    phones = re.findall(r"0\d{1,2}[- )]\d{3,4}[- ]\d{4}", normalized)
    mobile_phones = [phone for phone in phones if phone.startswith("010")]
    if mobile_phones:
        fields["contact"] = mobile_phones[-1]
    elif not fields["contact"] and phones:
        fields["contact"] = phones[-1]
    valid_dates = []
    for match in re.finditer(r"(?:(20\d{2})\s*(?:년|[-./]))?\s*(\d{1,2})\s*(?:월|[-./])\s*(\d{1,2})\s*(?:일)?", normalized):
        year, month, day = match.groups()
        if 1 <= int(month) <= 12 and 1 <= int(day) <= 31:
            valid_dates.append(f"{year + '-' if year else ''}{int(month):02d}-{int(day):02d}")
    if valid_dates:
        fields["request_date"] = valid_dates[-1]
    quantities = [int(value.replace(",", "")) for value in re.findall(r"(?<!\d)(\d{1,3}(?:,\d{3})*)\s*(?:개|EA|ea)", normalized)]
    if not fields["quantity"] and quantities:
        fields["quantity"] = str(max(quantities))
    if not fields["quantity"] and vendor == "고려기프트":
        candidates = [int(value.replace(",", "")) for value in re.findall(r"(?<![-\d])([1-9]\d{1,4})(?![-\d])", normalized)]
        repeated = [value for value in candidates if 1 <= value <= 10000 and candidates.count(value) >= 2]
        if repeated:
            fields["quantity"] = str(max(set(repeated), key=repeated.count))
    if not fields["product"]:
        product_match = re.search(r"([가-힣A-Za-z0-9+*() ]{8,}(?:배터리|충전기)[가-힣A-Za-z0-9+*() ]*)", normalized)
        if product_match:
            fields["product"] = _clean(product_match.group(1))[:100]
    if not fields["product"] and vendor == "고려기프트":
        product_match = re.search(r"([가-힣A-Za-z0-9+*() ]{0,35}보조.{0,8}(?:터리|Ei리).{0,35})", normalized)
        if product_match:
            fields["product"] = _clean(product_match.group(1))[:100]
    address_match = re.search(r"((?:서울|부산|대구|인천|광주|대전|울산|세종|경기|강원|충북|충남|전북|전남|경북|경남|제주)[^\n]{8,100})", normalized)
    if not fields["address"] and address_match:
        fields["address"] = _clean(address_match.group(1))
    if vendor == "고려기프트":
        if "선물포장" in normalized:
            fields["packaging"] = "선물포장"
        if "선불택" in normalized:
            fields["delivery"] = "택배"
        printing_terms = [term for term in ("컬러인쇄", "스티커", "선물포장") if term in normalized]
        if printing_terms:
            fields["printing"] = " + ".join(printing_terms)

    confidence = {key: (90 if value and source_type != "이미지 OCR" else 70 if value else 0) for key, value in fields.items()}
    return AnalysisResult(source_type, vendor, fields, confidence, normalized.strip())


def analyze_order_document(path: str | Path) -> AnalysisResult:
    text, source_type = extract_text(path)
    if not text.strip():
        raise ValueError("발주서에서 읽을 수 있는 텍스트가 없습니다.")
    return analyze_text(text, source_type)
