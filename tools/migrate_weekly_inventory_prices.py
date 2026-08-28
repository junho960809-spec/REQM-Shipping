"""기준 주간재고 Excel의 단가를 Supabase 전용 테이블로 1회 이관한다."""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

from openpyxl import load_workbook
from supabase import create_client

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from weekly_inventory_catalog import WEEKLY_INVENTORY_ITEMS


SUPABASE_URL = os.getenv("SUPABASE_URL", "https://jcslohuraqclhryeqxoc.supabase.co")
TABLE = "weekly_inventory_item_settings"


def workbook_rows(path: Path) -> list[dict]:
    active_codes = {code.casefold() for code, _ in WEEKLY_INVENTORY_ITEMS}
    active_order = {code.casefold(): index for index, (code, _) in enumerate(WEEKLY_INVENTORY_ITEMS)}
    fallback_names = {code.casefold(): name for code, name in WEEKLY_INVENTORY_ITEMS}
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = next((name for name in workbook.sheetnames if name.strip() == "단가"), None)
        if sheet_name is None:
            raise ValueError("단가 시트를 찾지 못했습니다.")
        rows: dict[str, dict] = {}
        for code, name, base_price, vat_price in workbook[sheet_name].iter_rows(min_row=3, max_col=4, values_only=True):
            item_code = str(code or "").strip()
            item_name = str(name or "").strip()
            if not item_code or not item_name or item_code in {"합계", "계"}:
                continue
            if base_price in (None, "") and vat_price not in (None, ""):
                base_price = float(vat_price or 0) / 1.1
            rows[item_code.casefold()] = {
                "item_code": item_code,
                "item_name": item_name,
                "base_unit_cost": float(base_price or 0),
                "is_active": item_code.casefold() in active_codes,
                "display_order": active_order.get(item_code.casefold(), 10000 + len(rows)),
            }
        for key in active_codes - rows.keys():
            code, name = next((c, n) for c, n in WEEKLY_INVENTORY_ITEMS if c.casefold() == key)
            rows[key] = {
                "item_code": code, "item_name": fallback_names[key], "base_unit_cost": 0,
                "is_active": True, "display_order": active_order[key],
            }
        return list(rows.values())
    finally:
        workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    args = parser.parse_args()
    key = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip()
    if not key:
        print("SUPABASE_SERVICE_ROLE_KEY 환경 변수가 필요합니다.", file=sys.stderr)
        return 2
    rows = workbook_rows(args.workbook)
    client = create_client(SUPABASE_URL, key)
    for start in range(0, len(rows), 500):
        client.table(TABLE).upsert(rows[start:start + 500], on_conflict="item_code").execute()
    print(f"단가 {len(rows):,}개 이관 · 주간재고 사용 {sum(row['is_active'] for row in rows):,}개")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
