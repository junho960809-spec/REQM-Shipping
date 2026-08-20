from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


SOURCE = Path(r"C:\Users\82104\OneDrive\Desktop\주간재고현황-20260814.xlsx")
OUTPUT = Path(__file__).resolve().parents[1] / "assets" / "weekly_inventory_template.xlsx"
ITEM_END_ROW = 169


def clear_values(sheet, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None


def replace_plain_sheet(workbook, name: str, headers: list[str]) -> None:
    index = workbook.sheetnames.index(name)
    workbook.remove(workbook[name])
    sheet = workbook.create_sheet(name, index)
    sheet.append(headers)
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False


def main() -> None:
    workbook = load_workbook(SOURCE, data_only=False)
    inventory = workbook["재고현황"]
    if inventory.max_row > ITEM_END_ROW:
        inventory.delete_rows(ITEM_END_ROW + 1, inventory.max_row - ITEM_END_ROW)
    clear_values(inventory, 5, ITEM_END_ROW, 6, 37)

    comparison = workbook["실재고 전산비교"]
    if comparison.max_row > ITEM_END_ROW:
        comparison.delete_rows(ITEM_END_ROW + 1, comparison.max_row - ITEM_END_ROW)
    clear_values(comparison, 5, ITEM_END_ROW, 4, 16)

    data = workbook["재고데이터"]
    clear_values(data, 3, data.max_row, 1, 12)
    reqm = workbook["재고데이터-리큐엠"]
    clear_values(reqm, 2, reqm.max_row, 1, 5)
    wekeep = workbook["재고데이터-위킵"]
    clear_values(wekeep, 2, wekeep.max_row, 1, 3)
    imoi = workbook["재고데이터-이모아이"]
    clear_values(imoi, 2, imoi.max_row, 1, 5)

    replace_plain_sheet(
        workbook,
        "RAWDATA_이카운트",
        ["기준일시", "창고코드", "창고명", "품목코드", "품목명", "현재고"],
    )
    replace_plain_sheet(workbook, "DAILY", ["현재 단계에서는 판매 데이터 제외"])
    replace_plain_sheet(workbook, "단가 ", ["현재 단계에서는 단가 데이터 제외"])

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    workbook.close()
    print(f"saved {OUTPUT} ({OUTPUT.stat().st_size:,} bytes)")


if __name__ == "__main__":
    sys.exit(main())
