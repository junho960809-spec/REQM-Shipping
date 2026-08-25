from __future__ import annotations

from copy import copy
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


def _comparison_text(value: str) -> str:
    without_notes = re.sub(r"\([^)]*\)", "", str(value or ""))
    return re.sub(r"[^0-9A-Za-z가-힣]", "", without_notes).casefold()


def clean_memo(reason: str, memo: str) -> str:
    """사유와 중복되는 메모 조각 및 구분 슬래시를 엑셀 비고에서 제거한다."""
    reason_key = _comparison_text(reason)
    remaining: list[str] = []
    for part in re.split(r"[/／]+", str(memo or "")):
        text = part.strip(" \t\r\n,;·-")
        if not text:
            continue
        part_key = _comparison_text(text)
        if reason_key and part_key == reason_key:
            continue
        remaining.append(text)
    return ", ".join(remaining)


def _copy_row_style(sheet, source: int, target: int, start_col: int = 2, end_col: int = 11) -> None:
    sheet.row_dimensions[target].height = sheet.row_dimensions[source].height
    for column in range(start_col, end_col + 1):
        src, dst = sheet.cell(source, column), sheet.cell(target, column)
        dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)


def _find_title_row(sheet, title: str) -> int | None:
    for row in range(1, sheet.max_row + 1):
        if any(str(sheet.cell(row, column).value or "").strip() == title for column in range(1, sheet.max_column + 1)):
            return row
    return None


def _blank_template():
    book = Workbook()
    sheet = book.active
    sheet.title = "송장양식"
    sections = [(1, "교환 출고 내역", ["이름", "우편번호", "주소", "연락처", "생산년월", "상품명", "수량", "사유", "비고", "송장번호"]),
                (16, "맞교환 입고 내역", ["이름", "맞교환 사유", "상품명", "입고일", "이상 유무", "생산년월", "확인 결과", "확인 결과", "확인 결과"]),
                (25, "반품 입고 내역", ["이름", "생산년월", "반품 사유", "상품명", "입고일", "이상 유무", "구매처", "제품가", "제품가", "제품가"])]
    for row, title, headers in sections:
        sheet.cell(row, 2, title).font = Font(bold=True, size=14)
        for offset, header in enumerate(headers, 2):
            cell = sheet.cell(row + 1, offset, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E8EFEF")
            cell.alignment = Alignment(horizontal="center")
    widths = {2: 13, 3: 14, 4: 45, 5: 18, 6: 13, 7: 25, 8: 10, 9: 28, 10: 22, 11: 18}
    for column, width in widths.items():
        sheet.column_dimensions[sheet.cell(1, column).column_letter].width = width
    return book


def export_as_daily(records: list[dict], output_path: str, template_path: str = "") -> str:
    template = Path(template_path) if template_path else None
    book = load_workbook(template) if template and template.is_file() else _blank_template()
    sheet = book["송장양식"] if "송장양식" in book.sheetnames else book.active

    exchange_title = _find_title_row(sheet, "교환 출고 내역") or 1
    matched_title = _find_title_row(sheet, "맞교환 입고 내역") or 16
    return_title = _find_title_row(sheet, "반품 입고 내역") or 25
    exchange_rows = [row for row in records if row.get("type") == "교환"]
    return_rows = [row for row in records if row.get("type") == "반품"]

    exchange_capacity = max(0, matched_title - (exchange_title + 2))
    if len(exchange_rows) > exchange_capacity:
        sheet.insert_rows(matched_title, len(exchange_rows) - exchange_capacity)
        matched_title = _find_title_row(sheet, "맞교환 입고 내역") or matched_title
        return_title = _find_title_row(sheet, "반품 입고 내역") or return_title

    def clear_between(start: int, end: int) -> None:
        for row in range(start + 2, end):
            for column in range(2, 12):
                sheet.cell(row, column).value = None

    clear_between(exchange_title, matched_title)
    clear_between(matched_title, return_title)
    for row in range(return_title + 2, sheet.max_row + 1):
        for column in range(2, 12):
            sheet.cell(row, column).value = None

    for index, record in enumerate(exchange_rows, exchange_title + 2):
        _copy_row_style(sheet, exchange_title + 2, index)
        product = " ".join(part for part in (record.get("product", ""), record.get("color", "")) if part).strip()
        values = [record.get("name", ""), record.get("postcode", ""), record.get("address", ""), record.get("phone", ""),
                  record.get("manufacture", ""), product, int(record.get("quantity") or 0), record.get("reason", ""),
                  clean_memo(record.get("reason", ""), record.get("memo", "")), record.get("invoice", "")]
        for column, value in enumerate(values, 2):
            sheet.cell(index, column).value = value

    for index, record in enumerate(return_rows, return_title + 2):
        _copy_row_style(sheet, return_title + 2, index)
        product = " ".join(part for part in (record.get("product", ""), record.get("color", "")) if part).strip()
        values = [record.get("name", ""), record.get("manufacture", "") or "확인불가", record.get("reason", ""), product,
                  record.get("receipt_date", ""), "", record.get("purchase_place", ""), "", "",
                  clean_memo(record.get("reason", ""), record.get("memo", ""))]
        for column, value in enumerate(values, 2):
            sheet.cell(index, column).value = value

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    book.save(output)
    return str(output)
