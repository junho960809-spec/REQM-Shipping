from __future__ import annotations

import os
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from as_daily_export import clean_memo, export_as_daily
from as_site_client import parse_html


class AsDailyTests(unittest.TestCase):
    def test_memo_removes_duplicate_reason_and_slashes(self) -> None:
        self.assertEqual(clean_memo("반응 없음", "반응없음 / 본사 출고 /"), "본사 출고")
        self.assertEqual(
            clean_memo("이어폰 충전 불량(모델 입력 필수)", "이어폰 충전 불량 /"),
            "",
        )
        self.assertEqual(
            clean_memo("출력 안됨(C 포트)", "출력 안됨(C 포트) / 본사 출고 / 양품화 완료"),
            "본사 출고, 양품화 완료",
        )

    def test_parser_keeps_detail_link_and_korean_text(self) -> None:
        root = parse_html('<table><tr><td><a href="passivedata1.view.php?cs_no=1">보기</a></td><td>홍길동</td></tr></table>')
        self.assertEqual(root.find_all("a")[0].attrs["href"], "passivedata1.view.php?cs_no=1")
        self.assertIn("홍길동", root.find_all("tr")[0].text())

    def test_export_writes_exchange_and_return_sections(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            template = os.path.join(folder, "template.xlsx")
            output = os.path.join(folder, "output.xlsx")
            book = Workbook()
            sheet = book.active
            sheet.title = "송장양식"
            sheet["B1"] = "교환 출고 내역"
            sheet["B16"] = "맞교환 입고 내역"
            sheet["B25"] = "반품 입고 내역"
            book.save(template)
            records = [
                {"type": "교환", "name": "교환고객", "postcode": "12345", "address": "서울", "phone": "010", "manufacture": "25.08", "product": "QP1000C", "color": "민트", "quantity": "1", "reason": "불량", "memo": "불량 / 본사 출고 /"},
                {"type": "반품", "name": "반품고객", "manufacture": "25.07", "reason": "변심", "product": "QP2000C", "receipt_date": "2026-08-19", "purchase_place": "스마트스토어"},
            ]
            export_as_daily(records, output, template)
            result = load_workbook(output).active
            self.assertEqual(result["B3"].value, "교환고객")
            self.assertEqual(result["G3"].value, "QP1000C 민트")
            self.assertEqual(result["J3"].value, "본사 출고")
            self.assertEqual(result["B27"].value, "반품고객")
            self.assertEqual(result["H27"].value, "스마트스토어")


if __name__ == "__main__":
    unittest.main()
