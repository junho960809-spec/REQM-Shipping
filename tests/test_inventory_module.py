import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import Workbook, load_workbook
from PySide6.QtWidgets import QApplication, QPushButton

from inventory_module import (
    InventoryDialog,
    InventoryRow,
    WeeklyEcountCredentialDialog,
    export_inventory_workbook,
    import_wekeep_rows,
    import_reference_workbook,
)
from weekly_inventory_store import load_item_prices, load_sales_rows, monthly_sales


class InventoryModuleTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.app = QApplication.instance() or QApplication([])

    def test_entry_rows_follow_reference_workbook_order(self):
        rows = InventoryDialog._initial_rows([{"item_code": "SHOULD-NOT-APPEAR", "item_name": "제외 품목"}])

        self.assertEqual(len(rows), 165)
        self.assertEqual((rows[0].item_code, rows[0].item_name), ("QWC-Q1500GR", "[리큐엠] QWC-Q1500 무선충전기 그레이"))
        self.assertEqual((rows[-1].item_code, rows[-1].item_name), ("QMA-Bubblepad-Cr", "리큐엠 맥세이프 액세서리-버블패드_소프트 크림"))
        self.assertNotIn("SHOULD-NOT-APPEAR", {row.item_code for row in rows})

    def test_entry_table_defaults_to_excel_source_order(self):
        dialog = InventoryDialog([])
        try:
            self.assertEqual(dialog.sort_filter.currentText(), "엑셀 원본 순서")
            self.assertEqual(dialog.entry_table.item(0, 0).text(), "QWC-Q1500GR")
            self.assertEqual(dialog.entry_table.item(dialog.entry_table.rowCount() - 1, 0).text(), "QMA-Bubblepad-Cr")
        finally:
            dialog.close()

    def test_weekly_inventory_uses_simplified_three_step_workflow(self):
        dialog = InventoryDialog([])
        try:
            self.assertEqual(dialog.tabs.count(), 3)
            self.assertEqual(
                [dialog.tabs.tabText(index) for index in range(dialog.tabs.count())],
                ["1  자료 준비", "2  실재고 입력", "3  결과 검토 · Excel"],
            )
            buttons = {button.text() for button in dialog.findChildren(QPushButton)}
            self.assertIn("전산재고 · RAWDATA 한 번에 최신화", buttons)
            self.assertNotIn("실재고 입력", buttons)
            self.assertNotIn("API 정보 입력/변경", buttons)
        finally:
            dialog.close()

    def test_combined_refresh_starts_rawdata_after_inventory_succeeds(self):
        dialog = InventoryDialog([])
        try:
            dialog.combined_sync_active = True
            with patch.object(dialog, "sync_sales_rawdata") as sync_sales:
                dialog.on_ecount_loaded([])
            sync_sales.assert_called_once_with()
            self.assertIn("전산재고 최신화 완료", dialog.inventory_prep_status.text())
            self.assertIn("판매 RAWDATA 최신화 중", dialog.sales_prep_status.text())
        finally:
            dialog.close()

    def test_entry_edit_updates_only_the_changed_row(self):
        dialog = InventoryDialog([])
        try:
            with patch.object(dialog, "refresh_entry_table") as full_entry_refresh, patch.object(dialog, "refresh_review_table") as review_refresh:
                dialog.entry_table.item(0, 2).setText("10")

            full_entry_refresh.assert_not_called()
            review_refresh.assert_not_called()
            self.assertEqual(dialog.entry_table.item(0, 4).text(), "+10")
            self.assertEqual(dialog.entry_table.item(0, 8).text(), "+10")
            self.assertEqual(dialog.entry_table.item(0, dialog.entry_table.columnCount() - 1).text(), "차이")
        finally:
            dialog.close()

    def test_entry_table_supports_single_click_and_any_key_editing(self):
        dialog = InventoryDialog([])
        try:
            triggers = dialog.entry_table.editTriggers()
            self.assertTrue(triggers & dialog.entry_table.EditTrigger.SelectedClicked)
            self.assertTrue(triggers & dialog.entry_table.EditTrigger.AnyKeyPressed)
        finally:
            dialog.close()

    def test_ecount_rows_map_exact_warehouse_codes_to_weekly_items(self):
        dialog = InventoryDialog([])
        try:
            matched = dialog.apply_ecount_rows([
                {"code": "QWC-Q1500GR", "warehouse_code": "100", "stock": 12},
                {"code": "QWC-Q1500GR", "warehouse_code": "300", "stock": 7},
                {"code": "QWC-Q1500GR", "warehouse_code": "CS001", "stock": 99},
                {"code": "NOT-IN-WEEKLY-LIST", "warehouse_code": "100", "stock": 5},
            ])

            first = dialog.rows[0]
            self.assertEqual(matched, 1)
            self.assertEqual(first.ecount_headquarters, 12)
            self.assertEqual(first.ecount_wekeep, 7)
        finally:
            dialog.close()

    def test_weekly_dialog_saves_credentials_to_shared_stores(self):
        with patch("inventory_module.load_ecount_users", return_value=[]), patch("inventory_module.upsert_ecount_user") as save_user, patch("inventory_module.save_api_key") as save_key:
            dialog = WeeklyEcountCredentialDialog()
            dialog.user_id.setEditText("JUNHO191")
            dialog.employee_code.setText("EMP01")
            dialog.api_key.setText("secret")
            dialog.save_and_accept()

        save_user.assert_called_once()
        save_key.assert_called_once_with("JUNHO191", "secret")
        self.assertEqual(dialog.credentials["company_code"], "304293")
        self.assertEqual(dialog.credentials["zone"], "AB")
        dialog.close()

    def test_differences(self):
        row = InventoryRow("A", "품목", 10, 7, 5, 8)
        self.assertEqual(row.headquarters_difference, 3)
        self.assertEqual(row.wekeep_difference, -3)
        self.assertEqual(row.total_difference, 0)

    def test_import_wekeep_rows(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "wekeep.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "재고데이터-위킵"
            sheet.append(["상품관리코드", "상품명", "시점재고"])
            sheet.append(["QWC-Q1500GR", "무선충전기 그레이", 536])
            workbook.save(path)
            result = import_wekeep_rows(path)
        self.assertEqual(result["QWC-Q1500GR"], ("무선충전기 그레이", 536))

    def test_export_inventory_workbook(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "inventory.xlsx"
            export_inventory_workbook(path, [InventoryRow("QWC-Q1500GR", "품목", 10, 7, 5, 8)], raw_sales_rows=[])
            workbook = load_workbook(path, data_only=False)
            self.assertEqual(
                workbook.sheetnames,
                ["재고현황", "실재고 전산비교", "재고데이터", "재고데이터-리큐엠", "재고데이터-위킵", "재고데이터-이모아이", "RAWDATA_이카운트", "DAILY", "단가 "],
            )
            self.assertEqual(workbook["재고현황"]["F5"].value, 10)
            self.assertEqual(workbook["재고현황"]["G5"].value, 5)
            self.assertEqual(workbook["재고현황"]["H5"].value, "=SUM(F5:G5)")
            self.assertEqual(workbook["실재고 전산비교"]["D5"].value, 10)
            self.assertEqual(workbook["실재고 전산비교"]["E5"].value, 7)
            self.assertEqual(workbook["실재고 전산비교"]["F5"].value, "=D5-E5")
            self.assertEqual(workbook["실재고 전산비교"]["L5"].value, "=J5-K5")
            self.assertEqual(workbook["재고현황"]["I5"].value, "=IFERROR(VLOOKUP($C5,'단가 '!$A$3:$D$3,4,0),0)")
            self.assertEqual(workbook["단가 "]["D3"].value, 0)
            self.assertEqual(str(workbook["재고현황"].freeze_panes), "A5")
            self.assertIn("B1:C2", {str(value) for value in workbook["재고현황"].merged_cells.ranges})
            workbook.close()

    def test_reference_workbook_accumulates_without_duplicate_rows(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.xlsx"
            store = Path(temp_dir) / "store.sqlite3"
            workbook = Workbook()
            raw = workbook.active
            raw.title = "RAWDATA_이카운트"
            raw.append([None] * 16)
            raw.append(["년", "월", "일", "요일", "주차", "구간", "구분", "일별", "거래처코드", "거래처", "품목코드", "품목명", "수량", "공급가액", "부가세", "합계"])
            raw.append([2026, 8, 3, "월", "1W", "월초", "리큐엠", 20260803, "C1", "거래처", "QWC-Q1500GR", "품목", 3, 1000, 100, 1100])
            raw.append([2026, 8, 3, "월", "1W", "월초", "리큐엠", 20260803, "C1", "거래처", "QWC-Q1500GR", "품목", 3, 1000, 100, 1100])
            price = workbook.create_sheet("단가 ")
            price.append(["회사명"])
            price.append(["품목코드", "품목명", "재고단가", "재고단가(+v)"])
            price.append(["QWC-Q1500GR", "품목", 1000, 1100])
            workbook.save(source)
            first = import_reference_workbook(source, store)
            second = import_reference_workbook(source, store)

            self.assertEqual((first["inserted"], first["duplicates"]), (2, 0))
            self.assertEqual((second["inserted"], second["duplicates"]), (0, 2))
            self.assertEqual(len(load_sales_rows(store)), 2)
            self.assertEqual(monthly_sales([(2026, 8)], store)["qwc-q1500gr"][(2026, 8)], 6)
            self.assertEqual(load_item_prices(store)["qwc-q1500gr"], 1100)


if __name__ == "__main__":
    unittest.main()
