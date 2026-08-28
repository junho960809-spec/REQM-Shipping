from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
import sys

from openpyxl import load_workbook
from weekly_inventory_catalog import WEEKLY_INVENTORY_ITEMS
from ecount_client import EcountClient
from ecount_credential_store import load_api_key, save_api_key
from ecount_user_store import load_ecount_users, upsert_ecount_user
from integration_credential_store import load_integration_credentials
from ecount_sales_sync import previous_inventory_week, sync_ecount_sales
from weekly_inventory_store import (
    RAW_HEADERS,
    add_sales_rows,
    load_item_prices,
    load_sales_rows,
    monthly_sales,
    recent_months,
    sales_row_count,
    save_item_prices,
)
from weekly_inventory_supabase import (
    fetch_monthly_sales as fetch_supabase_monthly_sales,
    fetch_rows as fetch_supabase_sales_rows,
    row_count as supabase_sales_row_count,
    upload_rows as upload_supabase_sales_rows,
)
from weekly_inventory_prices import active_items, fetch_price_settings, price_map
from PySide6.QtCore import QThread, QTimer, Signal, Qt
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QCheckBox,
    QAbstractItemView,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QSpinBox,
    QStyledItemDelegate,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)


class SelectAllEditDelegate(QStyledItemDelegate):
    def setEditorData(self, editor, index) -> None:
        super().setEditorData(editor, index)
        if isinstance(editor, QLineEdit):
            QTimer.singleShot(0, editor.selectAll)


class FastEntryTable(QTableWidget):
    def keyPressEvent(self, event) -> None:
        if event.key() in (Qt.Key.Key_Delete, Qt.Key.Key_Backspace) and self.state() != QAbstractItemView.State.EditingState:
            changed = False
            for item in self.selectedItems():
                if item.flags() & Qt.ItemFlag.ItemIsEditable:
                    item.setText("0")
                    changed = True
            if changed:
                event.accept()
                return
        super().keyPressEvent(event)


class WeeklyInventoryWorker(QThread):
    succeeded = Signal(object)
    failed = Signal(str)

    def __init__(self, credentials: dict):
        super().__init__()
        self.credentials = credentials

    def run(self) -> None:
        try:
            self.succeeded.emit(EcountClient(**self.credentials).get_inventory_by_location())
        except Exception as exc:
            self.failed.emit(str(exc))


class WeeklySalesSyncWorker(QThread):
    succeeded = Signal(object)
    failed = Signal(str)

    def __init__(self, client, credentials: dict):
        super().__init__()
        self.client = client
        self.credentials = credentials

    def run(self) -> None:
        try:
            self.succeeded.emit(sync_ecount_sales(self.client, self.credentials))
        except Exception as exc:
            self.failed.emit(str(exc))


class WeeklyEcountCredentialDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.credentials: dict | None = None
        self.profiles = load_ecount_users()
        self.setWindowTitle("주간 재고조사 · 이카운트 API 정보")
        self.setMinimumWidth(520)

        self.company_code = QLineEdit("304293")
        self.company_code.setReadOnly(True)
        self.user_id = QComboBox()
        self.user_id.setEditable(True)
        self.user_id.addItems([profile["user_id"] for profile in self.profiles])
        self.employee_code = QLineEdit()
        self.employee_code.setPlaceholderText("이카운트 담당자코드")
        self.api_key = QLineEdit()
        self.api_key.setEchoMode(QLineEdit.EchoMode.Password)
        self.api_key.setPlaceholderText("Windows 사용자 전용 암호화 저장")
        self.zone = QLineEdit("AB")
        self.zone.setReadOnly(True)

        form = QFormLayout()
        form.addRow("회사코드", self.company_code)
        form.addRow("사용자 ID", self.user_id)
        form.addRow("담당자코드", self.employee_code)
        form.addRow("API 인증키", self.api_key)
        form.addRow("존", self.zone)
        guide = QLabel("저장한 사용자와 인증키는 창고이동 화면에서도 동일하게 사용할 수 있습니다.")
        guide.setWordWrap(True)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        buttons.button(QDialogButtonBox.StandardButton.Save).setText("저장")
        buttons.accepted.connect(self.save_and_accept)
        buttons.rejected.connect(self.reject)
        layout = QVBoxLayout(self)
        layout.addWidget(guide)
        layout.addLayout(form)
        layout.addWidget(buttons)
        self.user_id.currentTextChanged.connect(self.load_selected_user)
        self.load_selected_user(self.user_id.currentText())

    def load_selected_user(self, user_id: str) -> None:
        profile = next((row for row in self.profiles if row["user_id"].casefold() == user_id.strip().casefold()), None)
        self.employee_code.setText(profile["employee_code"] if profile else "")
        self.api_key.setText(load_api_key(user_id))

    def save_and_accept(self) -> None:
        user_id = self.user_id.currentText().strip()
        employee_code = self.employee_code.text().strip()
        api_key = self.api_key.text().strip()
        if not user_id or not employee_code or not api_key:
            QMessageBox.warning(self, "입력 확인", "사용자 ID, 담당자코드, API 인증키를 모두 입력하세요.")
            return
        try:
            upsert_ecount_user({"user_id": user_id, "employee_code": employee_code, "display_name": ""})
            save_api_key(user_id, api_key)
        except Exception as exc:
            QMessageBox.warning(self, "API 정보 저장 실패", str(exc))
            return
        self.credentials = {
            "company_code": self.company_code.text().strip(),
            "user_id": user_id,
            "api_key": api_key,
            "zone": self.zone.text().strip(),
            "test_mode": False,
        }
        self.accept()


@dataclass
class InventoryRow:
    item_code: str
    item_name: str
    headquarters_actual: int = 0
    ecount_headquarters: int = 0
    wekeep_actual: int = 0
    ecount_wekeep: int = 0
    reason: str = ""
    action: str = ""
    reviewed: bool = False
    unit_price: float = 0
    sales_by_month: dict[tuple[int, int], float] = field(default_factory=dict)

    @property
    def headquarters_difference(self) -> int:
        return self.headquarters_actual - self.ecount_headquarters

    @property
    def wekeep_difference(self) -> int:
        return self.wekeep_actual - self.ecount_wekeep

    @property
    def total_difference(self) -> int:
        return self.headquarters_difference + self.wekeep_difference


SAMPLE_ROWS = [
    InventoryRow("QWC-Q1500GR", "[리큐엠] QWC-Q1500 무선충전기 그레이", 8, 7, 536, 540),
    InventoryRow("QWC-Q1500WH", "[리큐엠] QWC-Q1500 무선충전기 화이트", 11, 11, 412, 410),
    InventoryRow("QWC-Q1500-GR-PM", "[리큐엠] QWC-Q1500 무선충전기 그레이-판촉용", 629, 630, 0, 0),
    InventoryRow("QWC-Q1500-WH-PM", "[리큐엠] QWC-Q1500 무선충전기 화이트-판촉용", 63, 60, 0, 0),
    InventoryRow("QWC-Q1500PK", "[리큐엠] QWC-Q1500 무선충전기 핑크", 15, 15, 128, 130),
    InventoryRow("QWC-Q3100S-WH", "[리큐엠] 3in1 무선충전기 Q3100S 화이트", 22, 20, 74, 74),
]


def import_wekeep_rows(path: str | Path) -> dict[str, tuple[str, int]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["재고데이터-위킵"] if "재고데이터-위킵" in workbook.sheetnames else workbook.active
        values = sheet.iter_rows(values_only=True)
        header = next(values, None)
        if not header:
            raise ValueError("위킵 재고 파일이 비어 있습니다.")
        normalized = [str(value or "").replace(" ", "").lower() for value in header]

        def find_column(*aliases: str) -> int | None:
            for alias in aliases:
                key = alias.replace(" ", "").lower()
                if key in normalized:
                    return normalized.index(key)
            return None

        code_column = find_column("상품관리코드", "품목코드", "상품코드")
        name_column = find_column("상품명", "품목명")
        quantity_column = find_column("시점재고", "재고수량", "현재고", "수량")
        if code_column is None or quantity_column is None:
            raise ValueError("상품관리코드와 시점재고 열을 찾지 못했습니다.")
        result: dict[str, tuple[str, int]] = {}
        for row in values:
            code = str(row[code_column] or "").strip() if code_column < len(row) else ""
            if not code:
                continue
            name = str(row[name_column] or "").strip() if name_column is not None and name_column < len(row) else ""
            raw_quantity = row[quantity_column] if quantity_column < len(row) else 0
            try:
                quantity = int(float(raw_quantity or 0))
            except (TypeError, ValueError) as exc:
                raise ValueError(f"{code}의 시점재고 값이 숫자가 아닙니다.") from exc
            result[code] = (name, quantity)
        return result
    finally:
        workbook.close()


def import_reference_workbook(
    path: str | Path,
    store_path: str | Path | None = None,
    supabase_client=None,
) -> dict[str, int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        raw_name = next((name for name in workbook.sheetnames if name.strip() == "RAWDATA_이카운트"), None)
        price_name = next((name for name in workbook.sheetnames if name.strip() == "단가"), None)
        if raw_name is None or price_name is None:
            raise ValueError("RAWDATA_이카운트와 단가 시트를 모두 포함한 Excel을 선택하세요.")
        raw_sheet = workbook[raw_name]
        raw_rows = raw_sheet.iter_rows(min_row=3, max_col=16, values_only=True)
        if supabase_client is not None:
            sales_result = upload_supabase_sales_rows(supabase_client, raw_rows)
            inserted, duplicates = sales_result["inserted"], sales_result["duplicates"]
        else:
            inserted, duplicates = add_sales_rows(raw_rows, store_path)
        price_sheet = workbook[price_name]
        prices = []
        for values in price_sheet.iter_rows(min_row=3, max_col=4, values_only=True):
            code, name, base_price, vat_price = values
            if not str(code or "").strip():
                continue
            price = vat_price
            if price in (None, "") and base_price not in (None, ""):
                try:
                    price = float(base_price) * 1.1
                except (TypeError, ValueError):
                    price = 0
            prices.append((str(code).strip(), str(name or "").strip(), float(price or 0)))
        price_count = save_item_prices(prices, store_path)
        return {
            "inserted": inserted,
            "duplicates": duplicates,
            "prices": price_count,
            "storage": "supabase" if supabase_client is not None else "local",
        }
    finally:
        workbook.close()


def weekly_template_path() -> Path:
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    return base / "assets" / "weekly_inventory_template.xlsx"


def export_inventory_workbook(
    path: str | Path,
    rows: list[InventoryRow],
    template_path: str | Path | None = None,
    raw_sales_rows: list[list] | None = None,
) -> None:
    template = Path(template_path) if template_path else weekly_template_path()
    if not template.exists():
        raise FileNotFoundError(f"주간 재고조사 Excel 템플릿을 찾을 수 없습니다: {template}")
    workbook = load_workbook(template, data_only=False)
    overview = workbook["재고현황"]
    comparison = workbook["실재고 전산비교"]
    inventory_data = workbook["재고데이터"]
    reqm_data = workbook["재고데이터-리큐엠"]
    wekeep_data = workbook["재고데이터-위킵"]
    raw_ecount = workbook["RAWDATA_이카운트"]
    price_sheet = workbook[next(name for name in workbook.sheetnames if name.strip() == "단가")]
    export_raw_rows = raw_sales_rows if raw_sales_rows is not None else load_sales_rows()
    raw_last_row = max(3, len(export_raw_rows) + 2)
    price_last_row = max(3, len(rows) + 2)

    overview_by_code = {
        str(overview.cell(row, 3).value or "").strip().casefold(): row
        for row in range(5, 170)
    }
    comparison_by_code = {
        str(comparison.cell(row, 3).value or "").strip().casefold(): row
        for row in range(5, 170)
    }
    months = recent_months()
    for offset, (year, month) in enumerate(months, 12):
        overview.cell(1, offset).value = year
        overview.cell(2, offset).value = month
    price_sheet.cell(1, 1).value = f"회사명 : 주식회사 리큐엠 / {datetime.now():%Y/%m/%d}"
    for column, value in enumerate(["품목코드", "품목명[규격]", "재고단가", "재고단가(+v)"], 1):
        price_sheet.cell(2, column).value = value
    for row_number in range(3, price_sheet.max_row + 1):
        for column in range(1, 5):
            price_sheet.cell(row_number, column).value = None
    for index, item in enumerate(rows):
        overview_row = overview_by_code.get(item.item_code.casefold())
        if overview_row:
            overview.cell(overview_row, 6).value = item.headquarters_actual
            overview.cell(overview_row, 7).value = item.wekeep_actual
            overview.cell(overview_row, 8).value = f"=SUM(F{overview_row}:G{overview_row})"
            overview.cell(overview_row, 9).value = f"=IFERROR(VLOOKUP($C{overview_row},'단가 '!$A$3:$D${price_last_row},4,0),0)"
            overview.cell(overview_row, 10).value = f"=H{overview_row}*I{overview_row}"
            for column, (year, month) in enumerate(months, 12):
                overview.cell(overview_row, column).value = (
                    f'=SUMIFS(\'RAWDATA_이카운트\'!$M$3:$M${raw_last_row},'
                    f"'RAWDATA_이카운트'!$A$3:$A${raw_last_row},{year},"
                    f"'RAWDATA_이카운트'!$B$3:$B${raw_last_row},{month},"
                    f"'RAWDATA_이카운트'!$K$3:$K${raw_last_row},$C{overview_row})"
                )

        comparison_row = comparison_by_code.get(item.item_code.casefold())
        if comparison_row:
            values = {
                4: item.headquarters_actual,
                5: item.ecount_headquarters,
                6: f"=D{comparison_row}-E{comparison_row}",
                7: item.wekeep_actual,
                8: item.ecount_wekeep,
                9: f"=G{comparison_row}-H{comparison_row}",
                10: f"=D{comparison_row}+G{comparison_row}",
                11: f"=E{comparison_row}+H{comparison_row}",
                12: f"=J{comparison_row}-K{comparison_row}",
                13: " / ".join(value for value in (item.reason, item.action, "검토완료" if item.reviewed else "") if value),
                14: 0,
                15: 0,
                16: f"=O{comparison_row}-N{comparison_row}",
            }
            for column, value in values.items():
                comparison.cell(comparison_row, column).value = value

        data_row = index + 3
        data_values = [
            item.item_code, item.item_name, item.ecount_wekeep, item.ecount_headquarters, 0, None,
            0, 0, 0, f"=G{data_row}+C{data_row}", f"=H{data_row}+D{data_row}", f"=I{data_row}+E{data_row}",
        ]
        for column, value in enumerate(data_values, 1):
            inventory_data.cell(data_row, column).value = value

        reqm_row = index + 2
        for column, value in enumerate([item.item_code, item.item_name, item.ecount_wekeep, item.ecount_headquarters, 0], 1):
            reqm_data.cell(reqm_row, column).value = value
        for column, value in enumerate([item.item_code, item.item_name, item.wekeep_actual], 1):
            wekeep_data.cell(reqm_row, column).value = value

        price_row = index + 3
        price_sheet.cell(price_row, 1).value = item.item_code
        price_sheet.cell(price_row, 2).value = item.item_name
        price_sheet.cell(price_row, 3).value = item.unit_price / 1.1 if item.unit_price else 0
        price_sheet.cell(price_row, 4).value = item.unit_price

    if raw_ecount.max_row > 2:
        raw_ecount.delete_rows(3, raw_ecount.max_row - 2)
    for column in range(1, 17):
        raw_ecount.cell(1, column).value = None
        raw_ecount.cell(2, column).value = RAW_HEADERS[column - 1]
    for sales_row in export_raw_rows:
        raw_ecount.append(list(sales_row)[:16])

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.active = workbook.sheetnames.index("재고현황")
    workbook.save(path)
    workbook.close()


class InventoryDialog(QDialog):
    def __init__(self, catalog_items: list[dict] | None = None, parent=None):
        super().__init__(parent)
        self.main_window = parent
        self.ecount_worker: WeeklyInventoryWorker | None = None
        self.sales_sync_worker: WeeklySalesSyncWorker | None = None
        self.ecount_credentials_override: dict | None = None
        self.setWindowTitle("REQM 주간 재고조사")
        self.resize(1480, 900)
        self.price_settings = self._load_price_settings()
        self.rows = self._initial_rows(catalog_items or [], self.price_settings)
        self.sales_months = recent_months()
        self._load_sales_and_prices()
        self.filtered_indices: list[int] = []
        self.current_review_index: int | None = None
        self._build_ui()
        self.refresh_all()

    @staticmethod
    def _initial_rows(catalog_items: list[dict], settings: list[dict] | None = None) -> list[InventoryRow]:
        configured = active_items(settings or [])
        source = configured or WEEKLY_INVENTORY_ITEMS
        return [InventoryRow(item_code=code, item_name=name) for code, name in source]

    def _load_price_settings(self) -> list[dict]:
        client = self._supabase_client()
        if client is None:
            return []
        try:
            return fetch_price_settings(client)
        except Exception:
            return []

    def _load_sales_and_prices(self) -> None:
        self.price_settings = self._load_price_settings()
        prices = price_map(self.price_settings) if self.price_settings else load_item_prices()
        client = self._supabase_client()
        try:
            sales = fetch_supabase_monthly_sales(client, self.sales_months) if client else monthly_sales(self.sales_months)
            self.sales_storage = "Supabase" if client else "로컬"
        except Exception:
            sales = monthly_sales(self.sales_months)
            self.sales_storage = "로컬(오프라인)"
        for row in self.rows:
            key = row.item_code.casefold()
            row.unit_price = prices.get(key, 0)
            row.sales_by_month = sales.get(key, {})

    def _supabase_client(self):
        return getattr(self.main_window, "supabase_client", None) if self.main_window is not None else None

    def _sales_row_count(self) -> int:
        client = self._supabase_client()
        if client is not None:
            try:
                return supabase_sales_row_count(client)
            except Exception:
                pass
        return sales_row_count()

    def _sales_rows_for_export(self) -> list[list]:
        client = self._supabase_client()
        if client is not None:
            try:
                return fetch_supabase_sales_rows(client)
            except Exception as exc:
                raise RuntimeError(
                    "Supabase RAWDATA를 읽지 못했습니다. "
                    "20260828_ecount_sales_rawdata.sql 마이그레이션 적용 여부를 확인하세요."
                ) from exc
        return load_sales_rows()

    def _build_ui(self) -> None:
        self.setStyleSheet("""
            QDialog { background:#f5f7fa; color:#172f52; font-family:'맑은 고딕'; font-size:12px; }
            QLabel#title { font-size:26px; font-weight:800; color:#10294a; }
            QLabel#guide { color:#718096; }
            QFrame#card { background:white; border:1px solid #dce4ed; border-radius:12px; }
            QTabWidget::pane { border:0; }
            QTabBar::tab { background:#e9eef5; color:#50627a; padding:12px 24px; margin-right:3px; border-radius:7px; font-weight:700; }
            QTabBar::tab:selected { background:#087f78; color:white; }
            QLineEdit,QComboBox,QSpinBox,QTextEdit { background:white; border:1px solid #cfd9e5; border-radius:7px; padding:7px; }
            QPushButton { background:white; color:#29476b; border:1px solid #cbd7e5; border-radius:8px; padding:9px 15px; font-weight:700; }
            QPushButton:hover { background:#edf6f5; border-color:#0d9488; }
            QPushButton#primary { background:#087f78; color:white; border:none; }
            QPushButton#primary:hover { background:#066b66; }
            QPushButton:disabled { background:#edf0f4; color:#9ca7b5; }
            QTableWidget { background:white; alternate-background-color:white; gridline-color:#e7ebf0; border:1px solid #dbe3ec; }
            QHeaderView::section { background:#17365d; color:white; padding:9px; border:0; border-right:1px solid #365578; font-weight:700; }
        """)
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 20, 24, 20)
        header = QHBoxLayout()
        title = QLabel("주간 재고조사")
        title.setObjectName("title")
        header.addWidget(title)
        header.addStretch(1)
        self.source_status = QLabel("테스트 데이터 · 이카운트 API 연결 전")
        self.source_status.setObjectName("guide")
        header.addWidget(self.source_status)
        root.addLayout(header)
        guide = QLabel("이카운트 전산재고와 본사·위킵 실재고를 비교하고 주간재고조사 Excel을 생성합니다.")
        guide.setObjectName("guide")
        root.addWidget(guide)
        self.tabs = QTabWidget()
        self.tabs.addTab(self._dashboard_tab(), "1  조사 준비")
        self.tabs.addTab(self._entry_tab(), "2  실재고 입력")
        self.tabs.addTab(self._import_tab(), "3  자료 최신화")
        self.tabs.addTab(self._review_tab(), "4  결과 검토")
        self.tabs.currentChanged.connect(self.on_tab_changed)
        root.addWidget(self.tabs, 1)

    def _card(self, title_text: str, body_text: str, button_text: str, callback, enabled: bool = True) -> QFrame:
        card = QFrame()
        card.setObjectName("card")
        layout = QVBoxLayout(card)
        title = QLabel(title_text)
        title.setStyleSheet("font-size:17px;font-weight:800;color:#17365d")
        body = QLabel(body_text)
        body.setWordWrap(True)
        body.setObjectName("guide")
        button = QPushButton(button_text)
        button.setObjectName("primary" if enabled else "")
        button.setEnabled(enabled)
        button.clicked.connect(callback)
        layout.addWidget(title)
        layout.addWidget(body)
        layout.addStretch(1)
        layout.addWidget(button)
        return card

    def _dashboard_tab(self) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout(widget)
        cards = QHBoxLayout()
        cards.addWidget(self._card("1. 이카운트 최신화", "저장된 인증정보로 본사·위킵 창고의 전산재고를 가져옵니다.", "이카운트 최신화", self.refresh_ecount))
        cards.addWidget(self._card("2. 위킵 Excel 불러오기", "상품관리코드와 시점재고를 자동으로 반영합니다.", "Excel 파일 선택", self.load_wekeep))
        cards.addWidget(self._card("3. 본사 실재고 입력", "입력하는 즉시 본사·전체 차이를 다시 계산합니다.", "실재고 입력", lambda: self.tabs.setCurrentIndex(1)))
        cards.addWidget(self._card("4. 결과 Excel 생성", "검토 결과를 주간재고조사 파일로 저장합니다.", "결과 Excel 생성", self.export_result))
        layout.addLayout(cards)
        self.dashboard_status = QLabel()
        self.dashboard_status.setObjectName("guide")
        self.dashboard_status.setStyleSheet("background:white;border:1px solid #dce4ed;border-radius:10px;padding:18px")
        layout.addWidget(self.dashboard_status)
        layout.addStretch(1)
        return widget

    def _entry_tab(self) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout(widget)
        tools = QHBoxLayout()
        self.search = QLineEdit()
        self.search.setPlaceholderText("품목코드 또는 품목명 검색")
        self.status_filter = QComboBox()
        self.status_filter.addItems(["전체", "차이 있음", "일치"])
        self.location_filter = QComboBox()
        self.location_filter.addItems(["전체 위치", "본사 차이", "위킵 차이", "양쪽 차이"])
        self.sort_filter = QComboBox()
        self.sort_filter.addItems(["엑셀 원본 순서", "차이 큰 순", "품목코드 순", "품목명 순"])
        tools.addWidget(self.search, 2)
        tools.addWidget(self.status_filter)
        tools.addWidget(self.location_filter)
        tools.addWidget(self.sort_filter)
        layout.addLayout(tools)
        self.entry_summary = QLabel()
        self.entry_summary.setObjectName("guide")
        layout.addWidget(self.entry_summary)
        self.entry_table = FastEntryTable()
        headers = ["품목코드", "품목명", "본사 실재고", "이카운트 본사", "본사 차이", "위킵 재고", "이카운트 위킵", "위킵 차이", "전체 차이", "단가(+V)"]
        headers.extend(f"{year}년 {month}월 판매" for year, month in self.sales_months)
        headers.append("판정")
        self._prepare_table(self.entry_table, headers)
        self.entry_table.setItemDelegate(SelectAllEditDelegate(self.entry_table))
        self.entry_table.setEditTriggers(
            QAbstractItemView.EditTrigger.CurrentChanged
            | QAbstractItemView.EditTrigger.SelectedClicked
            | QAbstractItemView.EditTrigger.AnyKeyPressed
        )
        layout.addWidget(self.entry_table, 1)
        self.search.textChanged.connect(self.refresh_entry_table)
        self.status_filter.currentIndexChanged.connect(self.refresh_entry_table)
        self.location_filter.currentIndexChanged.connect(self.refresh_entry_table)
        self.sort_filter.currentIndexChanged.connect(self.refresh_entry_table)
        self.entry_table.cellClicked.connect(self.open_entry_editor)
        self.entry_table.cellChanged.connect(self.on_entry_changed)
        return widget

    def _import_tab(self) -> QWidget:
        widget = QWidget()
        layout = QHBoxLayout(widget)
        settings = self._card("이카운트 API 정보", "주간 재고조사 안에서 사용자 ID·담당자코드·API 인증키를 등록합니다.", "API 정보 입력/변경", self.open_ecount_settings)
        ecount = self._card("이카운트 API", "저장된 사용자와 API 키로 창고코드 100(본사)·300(위킵)의 최신 재고를 불러옵니다.", "이카운트 최신화", self.refresh_ecount)
        wekeep = self._card("위킵 Excel", "지원 열: 상품관리코드 / 상품명 / 시점재고\n현재 테스트 단계에서는 Excel 파일을 직접 불러옵니다.", "위킵 Excel 선택", self.load_wekeep)
        start_date, end_date = previous_inventory_week()
        sales_sync = self._card(
            "이카운트 판매 RAWDATA",
            f"판매현황을 {start_date:%Y-%m-%d}~{end_date:%Y-%m-%d} 기간으로 자동 조회하고 Supabase의 해당 기간을 갱신합니다.",
            "판매자료 자동 동기화",
            self.sync_sales_rawdata,
        )
        layout.addWidget(settings)
        layout.addWidget(ecount)
        layout.addWidget(wekeep)
        layout.addWidget(sales_sync)
        return widget

    def _review_tab(self) -> QWidget:
        widget = QWidget()
        layout = QHBoxLayout(widget)
        left = QVBoxLayout()
        self.review_table = QTableWidget()
        self._prepare_table(self.review_table, ["품목코드", "품목명", "본사 차이", "위킵 차이", "전체 차이", "처리 상태"])
        self.review_table.cellClicked.connect(self.select_review_row)
        left.addWidget(self.review_table)
        right_card = QFrame()
        right_card.setObjectName("card")
        right = QVBoxLayout(right_card)
        self.review_item = QLabel("차이 품목을 선택하세요")
        self.review_item.setStyleSheet("font-size:16px;font-weight:800")
        self.reason_combo = QComboBox()
        self.reason_combo.addItems(["미확인", "입출고 시점 차이", "전산 등록 오류", "파손·분실", "위킵 반영 지연", "기타"])
        self.action_edit = QTextEdit()
        self.action_edit.setPlaceholderText("조치 내용을 입력하세요")
        self.reviewer_edit = QLineEdit()
        self.reviewer_edit.setPlaceholderText("담당자")
        self.reviewed_check = QCheckBox("검토 완료")
        save_button = QPushButton("선택 품목 저장")
        save_button.clicked.connect(self.save_review)
        export_button = QPushButton("주간재고조사 Excel 생성")
        export_button.setObjectName("primary")
        export_button.clicked.connect(self.export_result)
        form = QFormLayout()
        form.addRow("차이 원인", self.reason_combo)
        form.addRow("조치 내용", self.action_edit)
        form.addRow("담당자", self.reviewer_edit)
        right.addWidget(self.review_item)
        right.addLayout(form)
        right.addWidget(self.reviewed_check)
        right.addWidget(save_button)
        right.addStretch(1)
        right.addWidget(export_button)
        layout.addLayout(left, 2)
        layout.addWidget(right_card, 1)
        return widget

    @staticmethod
    def _prepare_table(table: QTableWidget, headers: list[str]) -> None:
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.verticalHeader().setVisible(False)
        table.verticalHeader().setDefaultSectionSize(36)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setStretchLastSection(True)
        table.setAlternatingRowColors(False)

    def _visible_indices(self) -> list[int]:
        query = self.search.text().strip().lower() if hasattr(self, "search") else ""
        status = self.status_filter.currentText() if hasattr(self, "status_filter") else "전체"
        location = self.location_filter.currentText() if hasattr(self, "location_filter") else "전체 위치"
        indices = []
        for index, row in enumerate(self.rows):
            if query and query not in f"{row.item_code} {row.item_name}".lower():
                continue
            if status == "차이 있음" and row.total_difference == 0:
                continue
            if status == "일치" and row.total_difference != 0:
                continue
            if location == "본사 차이" and row.headquarters_difference == 0:
                continue
            if location == "위킵 차이" and row.wekeep_difference == 0:
                continue
            if location == "양쪽 차이" and (row.headquarters_difference == 0 or row.wekeep_difference == 0):
                continue
            indices.append(index)
        sort = self.sort_filter.currentText() if hasattr(self, "sort_filter") else "엑셀 원본 순서"
        if sort == "품목코드 순":
            indices.sort(key=lambda i: self.rows[i].item_code)
        elif sort == "품목명 순":
            indices.sort(key=lambda i: self.rows[i].item_name)
        elif sort == "차이 큰 순":
            indices.sort(key=lambda i: (-abs(self.rows[i].total_difference), self.rows[i].item_code))
        return indices

    def refresh_all(self) -> None:
        self.refresh_entry_table()
        self.refresh_review_table()
        self.update_dashboard_status()

    def update_dashboard_status(self) -> None:
        differences = sum(row.total_difference != 0 for row in self.rows)
        reviewed = sum(row.reviewed for row in self.rows if row.total_difference != 0)
        self.dashboard_status.setText(
            f"현재 품목 {len(self.rows):,}개 · 차이 품목 {differences:,}개 · 검토 완료 {reviewed:,}개\n"
            f"판매 RAWDATA 누적 {self._sales_row_count():,}행 · 저장소 {getattr(self, 'sales_storage', '로컬')} · "
            f"단가 저장소 {'Supabase' if self.price_settings else '로컬(이전 데이터)'}."
        )

    def refresh_entry_table(self) -> None:
        self.filtered_indices = self._visible_indices()
        self.entry_table.blockSignals(True)
        self.entry_table.setRowCount(len(self.filtered_indices))
        for table_row, source_index in enumerate(self.filtered_indices):
            row = self.rows[source_index]
            values = [row.item_code, row.item_name, row.headquarters_actual, row.ecount_headquarters, row.headquarters_difference, row.wekeep_actual, row.ecount_wekeep, row.wekeep_difference, row.total_difference, round(row.unit_price)]
            values.extend(round(row.sales_by_month.get(month, 0)) for month in self.sales_months)
            values.append("일치" if row.total_difference == 0 else "차이")
            for column, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                item.setData(Qt.ItemDataRole.UserRole, source_index)
                if column not in (2, 5):
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if column in (4, 7, 8):
                    item.setFont(FontProxy.bold_font(item.font()))
                    item.setForeground(QColor("#064fbd") if int(value) > 0 else QColor("#d11f2f") if int(value) < 0 else QColor("#536174"))
                    item.setText(f"+{value}" if int(value) > 0 else str(value))
                item.setBackground(QColor("white"))
                self.entry_table.setItem(table_row, column, item)
        self.entry_table.blockSignals(False)
        self.entry_summary.setText(f"전체 {len(self.rows):,}개 중 {len(self.filtered_indices):,}개 표시 · 실재고 입력 즉시 차이가 계산됩니다.")

    def update_entry_row(self, table_row: int, source_index: int) -> None:
        row = self.rows[source_index]
        values = {
            2: row.headquarters_actual,
            4: row.headquarters_difference,
            5: row.wekeep_actual,
            7: row.wekeep_difference,
            8: row.total_difference,
            self.entry_table.columnCount() - 1: "일치" if row.total_difference == 0 else "차이",
        }
        self.entry_table.blockSignals(True)
        try:
            for column, value in values.items():
                item = self.entry_table.item(table_row, column)
                if item is None:
                    item = QTableWidgetItem()
                    self.entry_table.setItem(table_row, column, item)
                item.setData(Qt.ItemDataRole.UserRole, source_index)
                if column in (4, 7, 8):
                    numeric = int(value)
                    item.setText(f"+{numeric}" if numeric > 0 else str(numeric))
                    item.setFont(FontProxy.bold_font(item.font()))
                    item.setForeground(QColor("#064fbd") if numeric > 0 else QColor("#d11f2f") if numeric < 0 else QColor("#536174"))
                else:
                    item.setText(str(value))
                item.setBackground(QColor("white"))
        finally:
            self.entry_table.blockSignals(False)

    def open_entry_editor(self, table_row: int, column: int) -> None:
        if column in (2, 5):
            self.entry_table.editItem(self.entry_table.item(table_row, column))

    def on_entry_changed(self, table_row: int, column: int) -> None:
        if column not in (2, 5) or not (0 <= table_row < len(self.filtered_indices)):
            return
        source_index = self.filtered_indices[table_row]
        try:
            value = int(self.entry_table.item(table_row, column).text().replace(",", "") or 0)
        except ValueError:
            self.refresh_entry_table()
            return
        if column == 2:
            self.rows[source_index].headquarters_actual = value
        else:
            self.rows[source_index].wekeep_actual = value
        filtering_by_difference = self.status_filter.currentText() != "전체" or self.location_filter.currentText() != "전체 위치"
        if self.sort_filter.currentText() == "차이 큰 순" or filtering_by_difference:
            self.refresh_entry_table()
        else:
            self.update_entry_row(table_row, source_index)
        self.update_dashboard_status()

    def on_tab_changed(self, index: int) -> None:
        if index == 3:
            self.refresh_review_table()

    def refresh_ecount(self) -> None:
        if self.ecount_worker is not None and self.ecount_worker.isRunning():
            self.source_status.setText("이카운트 재고를 불러오는 중입니다...")
            return
        if self.main_window is None or not hasattr(self.main_window, "inventory_credentials"):
            QMessageBox.warning(self, "이카운트 연결", "이카운트 인증정보를 사용할 수 없습니다.")
            return
        try:
            credentials = self.ecount_credentials_override or self.main_window.inventory_credentials()
        except Exception as exc:
            if self.open_ecount_settings():
                self.refresh_ecount()
            else:
                self.source_status.setText(f"이카운트 API 정보가 필요합니다 · {exc}")
            return
        self.source_status.setText("이카운트 본사·위킵 재고를 불러오는 중...")
        self.ecount_worker = WeeklyInventoryWorker(credentials)
        self.ecount_worker.succeeded.connect(self.on_ecount_loaded)
        self.ecount_worker.failed.connect(self.on_ecount_failed)
        self.ecount_worker.finished.connect(self.release_ecount_worker)
        self.ecount_worker.start()

    def open_ecount_settings(self) -> bool:
        dialog = WeeklyEcountCredentialDialog(self)
        if dialog.exec() != QDialog.DialogCode.Accepted or not dialog.credentials:
            return False
        self.ecount_credentials_override = dialog.credentials
        self.source_status.setText(f"이카운트 API 정보 저장 완료 · 사용자 {dialog.credentials['user_id']}")
        return True

    def apply_ecount_rows(self, source_rows: list[dict]) -> int:
        targets = {row.item_code.casefold(): row for row in self.rows}
        for row in self.rows:
            row.ecount_headquarters = 0
            row.ecount_wekeep = 0
        matched_codes: set[str] = set()
        for source in source_rows:
            code = str(source.get("code") or "").strip()
            target = targets.get(code.casefold())
            if target is None:
                continue
            try:
                quantity = int(round(float(source.get("stock", 0) or 0)))
            except (TypeError, ValueError):
                quantity = 0
            warehouse_code = str(source.get("warehouse_code") or "").strip()
            if warehouse_code == "100":
                target.ecount_headquarters += quantity
                matched_codes.add(code.casefold())
            elif warehouse_code == "300":
                target.ecount_wekeep += quantity
                matched_codes.add(code.casefold())
        return len(matched_codes)

    def on_ecount_loaded(self, source_rows: list[dict]) -> None:
        matched = self.apply_ecount_rows(source_rows)
        self.source_status.setText(f"이카운트 최신화 완료 · {datetime.now():%Y-%m-%d %H:%M:%S} · {matched:,}개 품목")
        self.refresh_all()

    def on_ecount_failed(self, message: str) -> None:
        self.source_status.setText("이카운트 최신화 실패")
        QMessageBox.critical(self, "이카운트 최신화 실패", message)

    def release_ecount_worker(self) -> None:
        worker = self.ecount_worker
        self.ecount_worker = None
        if worker is not None:
            worker.deleteLater()

    def refresh_review_table(self) -> None:
        indices = [index for index, row in enumerate(self.rows) if row.total_difference != 0]
        indices.sort(key=lambda index: -abs(self.rows[index].total_difference))
        self.review_table.setRowCount(len(indices))
        for table_row, source_index in enumerate(indices):
            row = self.rows[source_index]
            values = [row.item_code, row.item_name, row.headquarters_difference, row.wekeep_difference, row.total_difference, "검토 완료" if row.reviewed else "미처리"]
            for column, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                item.setData(Qt.ItemDataRole.UserRole, source_index)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if column in (2, 3, 4):
                    numeric = int(value)
                    item.setText(f"+{numeric}" if numeric > 0 else str(numeric))
                    item.setForeground(QColor("#064fbd") if numeric > 0 else QColor("#d11f2f") if numeric < 0 else QColor("#536174"))
                item.setBackground(QColor("white"))
                self.review_table.setItem(table_row, column, item)

    def select_review_row(self, table_row: int, _column: int) -> None:
        item = self.review_table.item(table_row, 0)
        if not item:
            return
        self.current_review_index = int(item.data(Qt.ItemDataRole.UserRole))
        row = self.rows[self.current_review_index]
        self.review_item.setText(f"{row.item_code}\n{row.item_name}")
        reason_index = self.reason_combo.findText(row.reason or "미확인")
        self.reason_combo.setCurrentIndex(max(reason_index, 0))
        self.action_edit.setPlainText(row.action)
        self.reviewed_check.setChecked(row.reviewed)

    def save_review(self) -> None:
        if self.current_review_index is None:
            QMessageBox.information(self, "품목 선택", "왼쪽 표에서 검토할 품목을 선택하세요.")
            return
        row = self.rows[self.current_review_index]
        row.reason = self.reason_combo.currentText()
        row.action = self.action_edit.toPlainText().strip()
        row.reviewed = self.reviewed_check.isChecked()
        self.refresh_all()

    def load_wekeep(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "위킵 재고 Excel 선택", "", "Excel 파일 (*.xlsx *.xlsm)")
        if not path:
            return
        try:
            imported = import_wekeep_rows(path)
        except Exception as exc:
            QMessageBox.critical(self, "위킵 파일 오류", str(exc))
            return
        matched = 0
        for row in self.rows:
            if row.item_code in imported:
                name, quantity = imported[row.item_code]
                row.wekeep_actual = quantity
                if not row.item_name and name:
                    row.item_name = name
                matched += 1
        self.source_status.setText(f"위킵 반영 완료 · {Path(path).name} · {matched:,}개 품목")
        self.refresh_all()
        QMessageBox.information(self, "위킵 재고 반영", f"{len(imported):,}개 행을 읽고 현재 목록의 {matched:,}개 품목에 반영했습니다.")

    def sync_sales_rawdata(self) -> None:
        if self.sales_sync_worker is not None and self.sales_sync_worker.isRunning():
            self.source_status.setText("이카운트 판매자료를 동기화하는 중...")
            return
        client = self._supabase_client()
        if client is None:
            QMessageBox.warning(self, "판매자료 동기화", "Supabase에 로그인한 뒤 다시 실행하세요.")
            return
        credentials = load_integration_credentials()
        if not credentials.get("ecount_user_id") or not credentials.get("ecount_password"):
            QMessageBox.warning(self, "판매자료 동기화", "메인 화면의 연동 계정에서 이카운트 아이디와 비밀번호를 저장하세요.")
            return
        config = getattr(self.main_window, "inventory_credentials", lambda: {})()
        sync_credentials = {
            "company_code": config.get("company_code", "304293"),
            "user_id": credentials["ecount_user_id"],
            "password": credentials["ecount_password"],
        }
        start_date, end_date = previous_inventory_week()
        self.source_status.setText(f"이카운트 판매현황 조회 중 · {start_date:%Y-%m-%d}~{end_date:%Y-%m-%d}")
        self.sales_sync_worker = WeeklySalesSyncWorker(client, sync_credentials)
        self.sales_sync_worker.succeeded.connect(self.on_sales_sync_loaded)
        self.sales_sync_worker.failed.connect(self.on_sales_sync_failed)
        self.sales_sync_worker.finished.connect(self.release_sales_sync_worker)
        self.sales_sync_worker.start()

    def on_sales_sync_loaded(self, result: dict) -> None:
        self._load_sales_and_prices()
        self.refresh_all()
        self.source_status.setText(
            f"판매자료 동기화 완료 · {result['start_date']:%Y-%m-%d}~{result['end_date']:%Y-%m-%d} · "
            f"{int(result.get('row_count', 0)):,}행"
        )
        QMessageBox.information(
            self, "판매자료 동기화 완료",
            f"Supabase에 {int(result.get('row_count', 0)):,}행을 저장했습니다.\n"
            f"수량 {float(result.get('quantity_total', 0)):,.0f} · 공급가액 {int(result.get('supply_total', 0)):,}원",
        )

    def on_sales_sync_failed(self, message: str) -> None:
        self.source_status.setText("이카운트 판매자료 동기화 실패")
        QMessageBox.critical(self, "판매자료 동기화 실패", message)

    def release_sales_sync_worker(self) -> None:
        worker = self.sales_sync_worker
        self.sales_sync_worker = None
        if worker is not None:
            worker.deleteLater()

    def export_result(self) -> None:
        default_name = f"REQM_주간재고조사_{datetime.now():%Y%m%d}.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "주간재고조사 Excel 저장", default_name, "Excel 파일 (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            export_inventory_workbook(path, self.rows, raw_sales_rows=self._sales_rows_for_export())
        except Exception as exc:
            QMessageBox.critical(self, "Excel 생성 실패", str(exc))
            return
        QMessageBox.information(self, "Excel 생성 완료", f"주간재고조사 파일을 저장했습니다.\n{path}")


class FontProxy:
    @staticmethod
    def bold_font(font):
        font.setBold(True)
        return font
